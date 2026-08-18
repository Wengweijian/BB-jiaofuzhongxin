#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
贵阳8月 · 门店积分统计表 → 看板数据JSON
用法: python3 guiyang_parse.py <积分表.xlsx> [输出.json]
      python3 guiyang_parse.py <数据回传表.xlsx> [输出.json]  (自动识别业绩表)
"""
import openpyxl, json, sys
from collections import defaultdict
from datetime import datetime, timedelta

def is_perf_file(xlsx_path):
    """判断是否为业绩回传表（含'数据回传'或'达成进度'关键字）"""
    name = xlsx_path.lower()
    return ('回传' in name or '达成' in name or '进度公示' in name)

def is_rank_file(xlsx_path):
    """判断是否为业绩排行榜文件（含'排行榜'关键字）"""
    return '排行' in xlsx_path.lower()

def parse_rank(xlsx_path):
    """解析业绩排行榜：渠道排名/门店A组B组/军师长PK"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    from collections import defaultdict

    # ===== 数据表：渠道汇总+门店排行 =====
    channels, store_rank = [], []
    if '数据表' in wb.sheetnames:
        ws = wb['数据表']
        rows = list(ws.iter_rows(values_only=True))
        stores = []
        cur_region = cur_jun = cur_shi = cur_ch = None
        for r in rows[3:]:
            if not r or not any(v is not None for v in r[:6]):
                continue
            region = r[0] if r[0] else cur_region
            jun = r[1] if r[1] else cur_jun
            shi = r[2] if r[2] else cur_shi
            ch = r[3] if r[3] else cur_ch
            store = r[4]
            if not store:
                continue
            cur_region, cur_jun, cur_shi, cur_ch = region, jun, shi, ch
            stores.append({
                'region': region, 'jun': jun, 'shi': shi, 'ch': ch, 'store': store,
                'target': float(r[5]) if isinstance(r[5], (int, float)) else 0,
                'done': float(r[6]) if isinstance(r[6], (int, float)) else 0,
                'rate': float(r[7]) if isinstance(r[7], (int, float)) else 0,
            })
        ch_tot = defaultdict(lambda: [0, 0])
        for s in stores:
            ch_tot[s['ch']][0] += s['target']
            ch_tot[s['ch']][1] += s['done']
        for ch, (t, d) in ch_tot.items():
            channels.append({'name': ch, 'target': t, 'done': d, 'rate': d/t if t else 0})
        channels.sort(key=lambda x: -x['rate'])
        store_rank = [s for s in stores if s['target'] > 0 and s['store'] not in ('其他', '合计', '')]
        store_rank.sort(key=lambda x: -x['rate'])

    # ===== 门店A/B组排行榜 =====
    def read_group(sheet):
        arr = []
        if sheet not in wb.sheetnames:
            return arr
        for r in list(wb[sheet].iter_rows(values_only=True))[3:]:
            if r and r[4] and isinstance(r[4], str) and isinstance(r[5], (int, float)):
                arr.append({
                    'region': r[0], 'jun': r[1], 'shi': r[2], 'ch': r[3], 'store': r[4],
                    'target': float(r[5]), 'done': float(r[6]), 'rate': float(r[7]),
                    'rank': r[11], 'note': r[12] if len(r) > 12 else None
                })
        return arr

    groupA = read_group('门店A组排行榜')
    groupB = read_group('门店B组排行榜')

    # ===== 军师长PK =====
    junshi = []
    if '军师长PK排行榜' in wb.sheetnames:
        for r in list(wb['军师长PK排行榜'].iter_rows(values_only=True))[3:]:
            if r and r[3] and isinstance(r[3], str) and isinstance(r[4], (int, float)):
                junshi.append({
                    'region': r[0] or '', 'jun': r[1] or '', 'shi': r[2] or '', 'ch': r[3],
                    'target': float(r[4]), 'done': float(r[5]), 'rate': float(r[6]),
                    'rank': r[10] if len(r) > 10 else None
                })

    return {
        'kind': 'rank',
        'channels': channels,
        'store_rank': store_rank[:30],
        'store_bottom': store_rank[-15:][::-1],
        'groupA': groupA,
        'groupB': groupB,
        'junshi': junshi,
        'updated': datetime.now().strftime('%Y-%m-%d %H:%M'),
    }

def parse_perf(xlsx_path):
    """解析业绩回传表：目标/完成/完成率/差额/时间进度"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    # 找表头行（含'业绩目标'）
    header_idx = None
    for i, r in enumerate(rows):
        if r and any('业绩目标' in str(v) for v in r if v):
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("未找到业绩表头")
    header = rows[header_idx]
    idx = {str(v): j for j, v in enumerate(header) if v}
    def gi(r, key):
        j = idx.get(key)
        return r[j] if j is not None and j < len(r) else None

    perf_rows = []
    start_d = end_d = None
    for r in rows[header_idx+1:]:
        if not r or not any(v is not None for v in r):
            continue
        teacher = gi(r, '项目老师')
        region = gi(r, '负责区域/城市')
        target = gi(r, '业绩目标')
        done = gi(r, '完成')
        rate = gi(r, '完成率')
        diff = gi(r, '差额')
        if teacher is None and target is None:
            continue
        if not teacher and not region:
            continue
        perf_rows.append({
            'teacher': str(teacher) if teacher else '',
            'region': str(region) if region else '',
            'target': float(target) if isinstance(target, (int, float)) else 0,
            'done': float(done) if isinstance(done, (int, float)) else 0,
            'rate': float(rate) if isinstance(rate, (int, float)) else (float(done)/float(target) if target else 0),
            'diff': float(diff) if isinstance(diff, (int, float)) else 0,
        })
        # 记录项目起止
        sd = gi(r, '开始时间')
        ed = gi(r, '结束时间')
        if sd and isinstance(sd, (int, float)):
            start_d = sd
        if ed and isinstance(ed, (int, float)):
            end_d = ed

    # 时间进度
    base = datetime(1899, 12, 30)
    today = datetime.now()
    if start_d and end_d:
        s = base + timedelta(days=start_d)
        e = base + timedelta(days=end_d)
        total_days = (e - s).days or 1
        elapsed = max(0, min((today - s).days, total_days))
        time_pct = round(elapsed / total_days * 100, 1)
        remain_days = (e - today).days
    else:
        time_pct, remain_days = 0, 0

    total = perf_rows[-1] if perf_rows else None
    return {
        'rows': perf_rows,
        'time_pct': time_pct,
        'remain_days': max(0, remain_days),
        'total': total,
        'updated': today.strftime('%Y-%m-%d %H:%M'),
    }

def parse(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # ============ 每日存根：11+N天积分 ============
    ws = wb['每日存根']
    rows = list(ws.iter_rows(values_only=True))
    # 表头: 0区域 1军长 2师长 3渠道 4门店分组 5匹配词 6门店 7..17日期 18合计
    # 动态识别日期列
    header = rows[2]
    dates = []
    for v in header[7:19]:
        if isinstance(v, (int, float)) and v > 40000:
            dates.append(v)
    n_days = len(dates)
    date_labels = []
    from datetime import datetime, timedelta
    base = datetime(1899, 12, 30)
    for v in dates:
        date_labels.append((base + timedelta(days=v)).strftime('%m/%d'))

    stores_daily = []
    cur_region = cur_jun = cur_shi = cur_ch = cur_group = None
    for r in rows[3:]:
        if not r or not r[6] or not isinstance(r[6], str):
            continue
        region = r[0] if r[0] else cur_region
        jun = r[1] if r[1] else cur_jun
        shi = r[2] if r[2] else cur_shi
        ch = r[3] if r[3] else cur_ch
        group = r[4] if r[4] else cur_group
        cur_region, cur_jun, cur_shi, cur_ch, cur_group = region, jun, shi, ch, group
        daily = [v if isinstance(v, (int, float)) else 0 for v in r[7:7+n_days]]
        stores_daily.append({
            'region': region, 'jun': jun, 'shi': shi, 'ch': ch, 'group': group,
            'store': r[6], 'daily': daily, 'total': sum(daily)
        })

    # 区域每日汇总
    reg_daily = defaultdict(lambda: [0]*n_days)
    reg_cnt = defaultdict(int)
    for s in stores_daily:
        reg_cnt[s['region']] += 1
        for i, d in enumerate(s['daily']):
            reg_daily[s['region']][i] += d

    daily_total = [0]*n_days
    for s in stores_daily:
        for i, d in enumerate(s['daily']):
            daily_total[i] += d

    # ============ S/A/B组排名 ============
    ranks = {}
    for g in ['S组', 'A组', 'B组']:
        wsg = wb[g]
        arr = []
        for r in wsg.iter_rows(min_row=4, values_only=True):
            if r and r[6] and isinstance(r[6], str) and isinstance(r[17], (int, float)):
                arr.append({
                    'region': r[0], 'store': r[6], 'score': int(r[17]),
                    'daily_avg': r[18], 'rank': r[19],
                    'reward': r[20] if isinstance(r[20], (int, float)) else None
                })
        ranks[g] = arr

    # ============ 每日公示（最新一天） ============
    ACTIONS = ['晨会', '夕会', '每日一读', '上门量尺', 'KDS预约', '新增客资', '小红书截流', '捷报',
               '抖音宣发', '小红书发布', '老用户回访', '周末活动落地', '渠道活动落地']
    try:
        wst = wb['每日公示']
        today = []
        cur_region = None
        for r in list(wst.iter_rows(values_only=True))[5:]:
            if not r or not r[6] or not isinstance(r[6], str):
                continue
            region = r[0] if r[0] else cur_region
            cur_region = region
            acts = {}
            for i, name in enumerate(ACTIONS):
                v = r[7+i]
                acts[name] = v if isinstance(v, (int, float)) else 0
            total = r[20] if isinstance(r[20], (int, float)) else 0
            today.append({'region': region, 'store': r[6], 'acts': acts, 'total': total})
    except KeyError:
        # 无每日公示sheet则用最后一天存根
        today = []
        for s in stores_daily:
            acts = {a: 1 if s['daily'][-1] >= 5 else 0 for a in ACTIONS[:8]}
            today.append({'region': s['region'], 'store': s['store'], 'acts': acts, 'total': s['daily'][-1]})

    # 今日动作完成率
    act_done = defaultdict(int)
    for t in today:
        for a in ACTIONS[:8]:
            if t['acts'].get(a, 0) > 0:
                act_done[a] += 1

    # 区域当日达标
    reg_today = defaultdict(lambda: [0, 0])
    for t in today:
        reg_today[t['region']][1] += 1
        if t['total'] >= 30:
            reg_today[t['region']][0] += 1

    data = {
        'updated': datetime.now().strftime('%Y-%m-%d %H:%M'),
        'date_labels': date_labels,
        'n_days': n_days,
        'daily_total': daily_total,
        'reg_daily': {k: v for k, v in reg_daily.items()},
        'reg_cnt': dict(reg_cnt),
        'stores_daily': stores_daily,
        'ranks': ranks,
        'today': today,
        'act_done': {k: v for k, v in act_done.items()},
        'reg_today': {k: v for k, v in reg_today.items()},
        'total_stores': len(stores_daily),
        'total_score': sum(s['total'] for s in stores_daily),
    }
    return data

if __name__ == '__main__':
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else 'guiyang_data.json'
    if is_rank_file(src):
        data = parse_rank(src)
        with open(out, 'w') as f:
            json.dump(data, f, ensure_ascii=False)
        print(f"✅ 排行榜解析完成: {len(data['channels'])}个渠道 / A组{len(data['groupA'])}家 / B组{len(data['groupB'])}家 / 军师长PK {len(data['junshi'])}条")
        print(f"   输出: {out}")
    elif is_perf_file(src):
        data = parse_perf(src)
        data['kind'] = 'perf'
        with open(out, 'w') as f:
            json.dump(data, f, ensure_ascii=False)
        t = data['total'] or {}
        print(f"✅ 业绩表解析完成: 目标{round(t.get('target',0)/10000,1)}万 / 完成{round(t.get('done',0)/10000,1)}万 / 达成率{round(t.get('rate',0)*100,1)}%")
        print(f"   时间进度: {data['time_pct']}% · 剩余{data['remain_days']}天 · 输出: {out}")
    else:
        data = parse(src)
        data['kind'] = 'points'
        with open(out, 'w') as f:
            json.dump(data, f, ensure_ascii=False)
        print(f"✅ 积分表解析完成: {len(data['stores_daily'])}家门店, {data['n_days']}天, 总积分{data['total_score']}")
        print(f"   输出: {out}")
